import pandas as pd
import numpy as np
import re
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
import PyPDF2
import pdfplumber
from openpyxl import load_workbook
from pathlib import Path
import os
from werkzeug.utils import secure_filename
from sklearn.impute import KNNImputer
import warnings
warnings.filterwarnings('ignore')

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DocumentProcessor:
    """Handle document processing for financial statements"""
    
    ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls', 'csv', 'docx'}
    
    @staticmethod
    def allowed_file(filename: str) -> bool:
        """Check if file extension is allowed"""
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in DocumentProcessor.ALLOWED_EXTENSIONS
    
    @staticmethod
    def secure_filename_custom(filename: str) -> str:
        """Create secure filename"""
        return secure_filename(filename)
    
    @staticmethod
    def extract_text_from_pdf(file_path: str) -> str:
        """Extract text from PDF file"""
        try:
            text = ""
            with open(file_path, 'rb') as file:
                # Try pdfplumber first (better for tables)
                try:
                    with pdfplumber.open(file) as pdf:
                        for page in pdf.pages:
                            page_text = page.extract_text()
                            if page_text:
                                text += page_text + "\n"
                except Exception as e:
                    logger.warning(f"pdfplumber failed, trying PyPDF2: {e}")
                    # Fallback to PyPDF2
                    file.seek(0)
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        text += page.extract_text() + "\n"
            
            return text
        except Exception as e:
            logger.error(f"Error extracting text from PDF: {e}")
            return ""
    
    @staticmethod
    def extract_tables_from_pdf(file_path: str) -> List[pd.DataFrame]:
        """Extract tables from PDF file"""
        try:
            tables = []
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_tables = page.extract_tables()
                    for table in page_tables:
                        if table and len(table) > 1:  # Ensure table has data
                            df = pd.DataFrame(table[1:], columns=table[0])
                            tables.append(df)
            return tables
        except Exception as e:
            logger.error(f"Error extracting tables from PDF: {e}")
            return []
    
    @staticmethod
    def process_excel_file(file_path: str) -> Dict[str, pd.DataFrame]:
        """Process Excel file and return all sheets"""
        try:
            excel_data = {}
            workbook = load_workbook(file_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    excel_data[sheet_name] = df
                except Exception as e:
                    logger.warning(f"Error reading sheet {sheet_name}: {e}")
                    continue
            
            return excel_data
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            return {}
    
    @staticmethod
    def process_csv_file(file_path: str) -> pd.DataFrame:
        """Process CSV file"""
        try:
            # Try different encodings
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    return df
                except UnicodeDecodeError:
                    continue
            
            # If all encodings fail, try with error handling
            df = pd.read_csv(file_path, encoding='utf-8', errors='ignore')
            return df
            
        except Exception as e:
            logger.error(f"Error processing CSV file: {e}")
            return pd.DataFrame()

class FinancialDataExtractor:
    """Extract financial data from processed documents"""
    
    # Common financial statement keywords
    BALANCE_SHEET_KEYWORDS = [
        'balance sheet', 'statement of financial position', 'assets', 'liabilities', 'equity',
        'current assets', 'non-current assets', 'total assets', 'shareholders equity'
    ]
    
    CASH_FLOW_KEYWORDS = [
        'cash flow', 'statement of cash flows', 'operating activities', 'investing activities',
        'financing activities', 'net cash', 'cash and cash equivalents'
    ]
    
    INCOME_STATEMENT_KEYWORDS = [
        'income statement', 'profit and loss', 'revenue', 'net income', 'operating income',
        'earnings', 'comprehensive income'
    ]
    
    @staticmethod
    def identify_statement_type(text: str) -> str:
        """Identify the type of financial statement"""
        text_lower = text.lower()
        
        cash_flow_score = sum(1 for keyword in FinancialDataExtractor.CASH_FLOW_KEYWORDS 
                             if keyword in text_lower)
        balance_sheet_score = sum(1 for keyword in FinancialDataExtractor.BALANCE_SHEET_KEYWORDS 
                                 if keyword in text_lower)
        income_statement_score = sum(1 for keyword in FinancialDataExtractor.INCOME_STATEMENT_KEYWORDS 
                                   if keyword in text_lower)
        
        if cash_flow_score >= balance_sheet_score and cash_flow_score >= income_statement_score:
            return 'cash_flow'
        elif balance_sheet_score >= income_statement_score:
            return 'balance_sheet'
        else:
            return 'income_statement'
    
    @staticmethod
    def extract_financial_values(text: str) -> Dict[str, float]:
        """Extract financial values from text"""
        values = {}
        
        # Common financial line items patterns
        patterns = {
            'total_assets': r'total\s+assets.*?(\$?[\d,]+\.?\d*)',
            'current_assets': r'current\s+assets.*?(\$?[\d,]+\.?\d*)',
            'cash_and_equivalents': r'cash\s+and\s+cash\s+equivalents.*?(\$?[\d,]+\.?\d*)',
            'accounts_receivable': r'accounts\s+receivable.*?(\$?[\d,]+\.?\d*)',
            'inventory': r'inventory.*?(\$?[\d,]+\.?\d*)',
            'total_liabilities': r'total\s+liabilities.*?(\$?[\d,]+\.?\d*)',
            'current_liabilities': r'current\s+liabilities.*?(\$?[\d,]+\.?\d*)',
            'long_term_debt': r'long.?term\s+debt.*?(\$?[\d,]+\.?\d*)',
            'shareholders_equity': r'shareholders?\s+equity.*?(\$?[\d,]+\.?\d*)',
            'net_income': r'net\s+income.*?(\$?[\d,]+\.?\d*)',
            'revenue': r'revenue.*?(\$?[\d,]+\.?\d*)',
        }
        
        for key, pattern in patterns.items():
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                try:
                    # Clean and convert to float
                    value_str = matches[0].replace('$', '').replace(',', '')
                    values[key] = float(value_str)
                except ValueError:
                    continue
        
        return values
    
    @staticmethod
    def extract_year_from_text(text: str) -> int:
        """Extract year from financial statement"""
        current_year = datetime.now().year
        
        # Look for years in the text
        year_patterns = [
            r'\b(20\d{2})\b',  # 2000-2099
            r'\b(19\d{2})\b',  # 1900-1999
        ]
        
        years = []
        for pattern in year_patterns:
            matches = re.findall(pattern, text)
            years.extend([int(year) for year in matches])
        
        # Filter reasonable years (last 10 years to next 2 years)
        valid_years = [year for year in years 
                      if current_year - 10 <= year <= current_year + 2]
        
        if valid_years:
            return max(valid_years)  # Return most recent year
        
        return current_year
    
    @staticmethod
    def extract_balance_sheet_data(text: str, tables: List[pd.DataFrame] = None) -> Dict[str, Any]:
        """Extract comprehensive balance sheet data"""
        data = {}
        
        # Enhanced patterns for balance sheet items
        balance_sheet_patterns = {
            # Current Assets
            'cash_and_equivalents': [
                r'cash\s+and\s+cash\s+equivalents.*?(\$?[\d,]+\.?\d*)',
                r'cash\s+and\s+equivalents.*?(\$?[\d,]+\.?\d*)',
                r'cash.*?(\$?[\d,]+\.?\d*)'
            ],
            'accounts_receivable': [
                r'accounts\s+receivable.*?(\$?[\d,]+\.?\d*)',
                r'trade\s+receivables.*?(\$?[\d,]+\.?\d*)',
                r'receivables.*?(\$?[\d,]+\.?\d*)'
            ],
            'inventory': [
                r'inventory.*?(\$?[\d,]+\.?\d*)',
                r'inventories.*?(\$?[\d,]+\.?\d*)'
            ],
            'prepaid_expenses': [
                r'prepaid\s+expenses.*?(\$?[\d,]+\.?\d*)',
                r'prepaid.*?(\$?[\d,]+\.?\d*)'
            ],
            'current_assets': [
                r'total\s+current\s+assets.*?(\$?[\d,]+\.?\d*)',
                r'current\s+assets.*?(\$?[\d,]+\.?\d*)'
            ],
            
            # Non-Current Assets
            'property_plant_equipment': [
                r'property,?\s+plant\s+and\s+equipment.*?(\$?[\d,]+\.?\d*)',
                r'property,?\s+plant.*?(\$?[\d,]+\.?\d*)',
                r'fixed\s+assets.*?(\$?[\d,]+\.?\d*)'
            ],
            'intangible_assets': [
                r'intangible\s+assets.*?(\$?[\d,]+\.?\d*)',
                r'intangibles.*?(\$?[\d,]+\.?\d*)'
            ],
            'goodwill': [
                r'goodwill.*?(\$?[\d,]+\.?\d*)'
            ],
            'investments': [
                r'investments.*?(\$?[\d,]+\.?\d*)',
                r'investment\s+securities.*?(\$?[\d,]+\.?\d*)'
            ],
            'total_assets': [
                r'total\s+assets.*?(\$?[\d,]+\.?\d*)'
            ],
            
            # Current Liabilities
            'accounts_payable': [
                r'accounts\s+payable.*?(\$?[\d,]+\.?\d*)',
                r'trade\s+payables.*?(\$?[\d,]+\.?\d*)'
            ],
            'short_term_debt': [
                r'short.?term\s+debt.*?(\$?[\d,]+\.?\d*)',
                r'current\s+portion.*?debt.*?(\$?[\d,]+\.?\d*)'
            ],
            'accrued_liabilities': [
                r'accrued\s+liabilities.*?(\$?[\d,]+\.?\d*)',
                r'accrued\s+expenses.*?(\$?[\d,]+\.?\d*)'
            ],
            'current_liabilities': [
                r'total\s+current\s+liabilities.*?(\$?[\d,]+\.?\d*)',
                r'current\s+liabilities.*?(\$?[\d,]+\.?\d*)'
            ],
            
            # Non-Current Liabilities
            'long_term_debt': [
                r'long.?term\s+debt.*?(\$?[\d,]+\.?\d*)',
                r'non.?current\s+debt.*?(\$?[\d,]+\.?\d*)'
            ],
            'deferred_tax_liabilities': [
                r'deferred\s+tax\s+liabilities.*?(\$?[\d,]+\.?\d*)',
                r'deferred\s+taxes.*?(\$?[\d,]+\.?\d*)'
            ],
            'total_liabilities': [
                r'total\s+liabilities.*?(\$?[\d,]+\.?\d*)'
            ],
            
            # Equity
            'share_capital': [
                r'share\s+capital.*?(\$?[\d,]+\.?\d*)',
                r'common\s+stock.*?(\$?[\d,]+\.?\d*)',
                r'capital\s+stock.*?(\$?[\d,]+\.?\d*)'
            ],
            'retained_earnings': [
                r'retained\s+earnings.*?(\$?[\d,]+\.?\d*)'
            ],
            'total_equity': [
                r'total\s+shareholders?\s+equity.*?(\$?[\d,]+\.?\d*)',
                r'total\s+equity.*?(\$?[\d,]+\.?\d*)',
                r'shareholders?\s+equity.*?(\$?[\d,]+\.?\d*)'
            ]
        }
        
        # Extract values using patterns
        for field, patterns in balance_sheet_patterns.items():
            for pattern in patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                if matches:
                    try:
                        value_str = matches[0].replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
                        data[field] = float(value_str)
                        break  # Use first successful match
                    except ValueError:
                        continue
            
            # Set default value if not found
            if field not in data:
                data[field] = 0
        
        # Try to extract from tables if available
        if tables:
            data.update(FinancialDataExtractor._extract_from_tables(tables, 'balance_sheet'))
        
        return data
    
    @staticmethod
    def extract_cash_flow_data(text: str, tables: List[pd.DataFrame] = None) -> Dict[str, Any]:
        """Extract comprehensive cash flow data"""
        data = {}
        
        cash_flow_patterns = {
            'net_income': [
                r'net\s+income.*?(\$?[\d,]+\.?\d*)',
                r'net\s+earnings.*?(\$?[\d,]+\.?\d*)'
            ],
            'depreciation_and_amortization': [
                r'depreciation\s+and\s+amortization.*?(\$?[\d,]+\.?\d*)',
                r'depreciation.*?(\$?[\d,]+\.?\d*)'
            ],
            'stock_based_compensation': [
                r'stock.?based\s+compensation.*?(\$?[\d,]+\.?\d*)',
                r'share.?based\s+compensation.*?(\$?[\d,]+\.?\d*)'
            ],
            'changes_in_working_capital': [
                r'changes?\s+in\s+working\s+capital.*?(\$?[\d,]+\.?\d*)',
                r'working\s+capital\s+changes?.*?(\$?[\d,]+\.?\d*)'
            ],
            'net_cash_from_operating_activities': [
                r'net\s+cash\s+.*?operating\s+activities.*?(\$?[\d,]+\.?\d*)',
                r'cash\s+from\s+operating\s+activities.*?(\$?[\d,]+\.?\d*)'
            ],
            'capital_expenditures': [
                r'capital\s+expenditures.*?(\$?[\d,]+\.?\d*)',
                r'purchases?\s+of\s+.*?equipment.*?(\$?[\d,]+\.?\d*)'
            ],
            'acquisitions': [
                r'acquisitions.*?(\$?[\d,]+\.?\d*)',
                r'business\s+acquisitions.*?(\$?[\d,]+\.?\d*)'
            ],
            'net_cash_from_investing_activities': [
                r'net\s+cash\s+.*?investing\s+activities.*?(\$?[\d,]+\.?\d*)',
                r'cash\s+from\s+investing\s+activities.*?(\$?[\d,]+\.?\d*)'
            ],
            'dividends_paid': [
                r'dividends\s+paid.*?(\$?[\d,]+\.?\d*)',
                r'cash\s+dividends.*?(\$?[\d,]+\.?\d*)'
            ],
            'share_repurchases': [
                r'share\s+repurchases.*?(\$?[\d,]+\.?\d*)',
                r'repurchases?\s+of\s+.*?stock.*?(\$?[\d,]+\.?\d*)'
            ],
            'net_cash_from_financing_activities': [
                r'net\s+cash\s+.*?financing\s+activities.*?(\$?[\d,]+\.?\d*)',
                r'cash\s+from\s+financing\s+activities.*?(\$?[\d,]+\.?\d*)'
            ]
        }
        
        # Extract values using patterns
        for field, patterns in cash_flow_patterns.items():
            for pattern in patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                if matches:
                    try:
                        value_str = matches[0].replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
                        data[field] = float(value_str)
                        break
                    except ValueError:
                        continue
            
            if field not in data:
                data[field] = 0
        
        # Try to extract from tables if available
        if tables:
            data.update(FinancialDataExtractor._extract_from_tables(tables, 'cash_flow'))
        
        return data
    
    @staticmethod
    def _extract_from_tables(tables: List[pd.DataFrame], statement_type: str) -> Dict[str, float]:
        """Extract financial data from DataFrame tables"""
        extracted_data = {}
        
        for df in tables:
            if df.empty:
                continue
            
            try:
                # Convert dataframe to lowercase for pattern matching
                df_lower = df.astype(str).apply(lambda x: x.str.lower())
                
                # Look for financial keywords in the first column
                for idx, row in df_lower.iterrows():
                    first_col = str(row.iloc[0]) if len(row) > 0 else ""
                    
                    # Find numeric values in subsequent columns
                    for col_idx in range(1, len(row)):
                        cell_value = str(df.iloc[idx, col_idx])
                        
                        # Try to extract numeric value
                        numeric_match = re.search(r'(\$?[\d,]+\.?\d*)', cell_value)
                        if numeric_match:
                            try:
                                value = float(numeric_match.group(1).replace('$', '').replace(',', ''))
                                
                                # Map to appropriate field based on keywords
                                if statement_type == 'balance_sheet':
                                    field = FinancialDataExtractor._map_balance_sheet_field(first_col)
                                else:  # cash_flow
                                    field = FinancialDataExtractor._map_cash_flow_field(first_col)
                                
                                if field:
                                    extracted_data[field] = value
                                    break
                                    
                            except ValueError:
                                continue
                                
            except Exception as e:
                logger.warning(f"Error processing table: {e}")
                continue
        
        return extracted_data
    
    @staticmethod
    def _map_balance_sheet_field(text: str) -> Optional[str]:
        """Map text to balance sheet field names"""
        text = text.lower()
        
        field_mappings = {
            'cash': 'cash_and_equivalents',
            'receivable': 'accounts_receivable',
            'inventory': 'inventory',
            'current assets': 'current_assets',
            'total assets': 'total_assets',
            'payable': 'accounts_payable',
            'current liabilities': 'current_liabilities',
            'total liabilities': 'total_liabilities',
            'equity': 'total_equity',
            'retained earnings': 'retained_earnings'
        }
        
        for keyword, field in field_mappings.items():
            if keyword in text:
                return field
        
        return None
    
    @staticmethod
    def _map_cash_flow_field(text: str) -> Optional[str]:
        """Map text to cash flow field names"""
        text = text.lower()
        
        field_mappings = {
            'net income': 'net_income',
            'depreciation': 'depreciation_and_amortization',
            'operating activities': 'net_cash_from_operating_activities',
            'capital expenditure': 'capital_expenditures',
            'investing activities': 'net_cash_from_investing_activities',
            'financing activities': 'net_cash_from_financing_activities',
            'dividends': 'dividends_paid'
        }
        
        for keyword, field in field_mappings.items():
            if keyword in text:
                return field
        
        return None

class BalanceSheetProcessor:
    """Process balance sheet data and calculate ratios"""
    
    @staticmethod
    def process_balance_sheet_data(data: Dict[str, Any]) -> Dict[str, Any]:
        """Process and validate balance sheet data"""
        
        processed_data = {
            'current_assets': data.get('current_assets', 0),
            'cash_and_equivalents': data.get('cash_and_equivalents', 0),
            'accounts_receivable': data.get('accounts_receivable', 0),
            'inventory': data.get('inventory', 0),
            'prepaid_expenses': data.get('prepaid_expenses', 0),
            'other_current_assets': data.get('other_current_assets', 0),
            'non_current_assets': data.get('non_current_assets', 0),
            'property_plant_equipment': data.get('property_plant_equipment', 0),
            'accumulated_depreciation': data.get('accumulated_depreciation', 0),
            'net_ppe': data.get('net_ppe', 0),
            'intangible_assets': data.get('intangible_assets', 0),
            'goodwill': data.get('goodwill', 0),
            'investments': data.get('investments', 0),
            'other_non_current_assets': data.get('other_non_current_assets', 0),
            'total_assets': data.get('total_assets', 0),
            'current_liabilities': data.get('current_liabilities', 0),
            'accounts_payable': data.get('accounts_payable', 0),
            'short_term_debt': data.get('short_term_debt', 0),
            'accrued_liabilities': data.get('accrued_liabilities', 0),
            'deferred_revenue': data.get('deferred_revenue', 0),
            'other_current_liabilities': data.get('other_current_liabilities', 0),
            'non_current_liabilities': data.get('non_current_liabilities', 0),
            'long_term_debt': data.get('long_term_debt', 0),
            'deferred_tax_liabilities': data.get('deferred_tax_liabilities', 0),
            'pension_obligations': data.get('pension_obligations', 0),
            'other_non_current_liabilities': data.get('other_non_current_liabilities', 0),
            'total_liabilities': data.get('total_liabilities', 0),
            'share_capital': data.get('share_capital', 0),
            'retained_earnings': data.get('retained_earnings', 0),
            'additional_paid_in_capital': data.get('additional_paid_in_capital', 0),
            'treasury_stock': data.get('treasury_stock', 0),
            'accumulated_other_comprehensive_income': data.get('accumulated_other_comprehensive_income', 0),
            'total_equity': data.get('total_equity', 0),
        }
        
        # Calculate missing values if possible
        processed_data = BalanceSheetProcessor._calculate_missing_values(processed_data)
        
        # Validate balance sheet equation
        processed_data['balance_check'] = BalanceSheetProcessor._validate_balance_equation(processed_data)
        
        # Calculate accuracy percentage
        processed_data['accuracy_percentage'] = BalanceSheetProcessor._calculate_accuracy(processed_data)
        
        return processed_data
    
    @staticmethod
    def _calculate_missing_values(data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate missing values using financial relationships"""
        
        # Calculate total current assets if missing
        if data['current_assets'] == 0:
            data['current_assets'] = (
                data['cash_and_equivalents'] + 
                data['accounts_receivable'] + 
                data['inventory'] + 
                data['prepaid_expenses'] + 
                data['other_current_assets']
            )
        
        # Calculate net PPE if missing
        if data['net_ppe'] == 0 and data['property_plant_equipment'] > 0:
            data['net_ppe'] = data['property_plant_equipment'] - data['accumulated_depreciation']
        elif data['net_ppe'] == 0:
            data['net_ppe'] = data['property_plant_equipment']
        
        # Calculate total non-current assets if missing
        if data['non_current_assets'] == 0:
            data['non_current_assets'] = (
                data['net_ppe'] + 
                data['intangible_assets'] + 
                data['goodwill'] + 
                data['investments'] + 
                data['other_non_current_assets']
            )
        
        # Calculate total assets if missing
        if data['total_assets'] == 0:
            data['total_assets'] = data['current_assets'] + data['non_current_assets']
        
        # Calculate total current liabilities if missing
        if data['current_liabilities'] == 0:
            data['current_liabilities'] = (
                data['accounts_payable'] + 
                data['short_term_debt'] + 
                data['accrued_liabilities'] + 
                data['deferred_revenue'] + 
                data['other_current_liabilities']
            )
        
        # Calculate total non-current liabilities if missing
        if data['non_current_liabilities'] == 0:
            data['non_current_liabilities'] = (
                data['long_term_debt'] + 
                data['deferred_tax_liabilities'] + 
                data['pension_obligations'] + 
                data['other_non_current_liabilities']
            )
        
        # Calculate total liabilities if missing
        if data['total_liabilities'] == 0:
            data['total_liabilities'] = data['current_liabilities'] + data['non_current_liabilities']
        
        # Calculate total equity if missing
        if data['total_equity'] == 0:
            data['total_equity'] = (
                data['share_capital'] + 
                data['retained_earnings'] + 
                data['additional_paid_in_capital'] - 
                data['treasury_stock'] + 
                data['accumulated_other_comprehensive_income']
            )
        
        # Final balance check - calculate equity from assets and liabilities if needed
        if data['total_equity'] == 0 and data['total_assets'] > 0:
            data['total_equity'] = data['total_assets'] - data['total_liabilities']
        
        return data
    
    @staticmethod
    def _validate_balance_equation(data: Dict[str, Any]) -> bool:
        """Validate that Assets = Liabilities + Equity"""
        total_assets = data['total_assets']
        total_liab_equity = data['total_liabilities'] + data['total_equity']
        
        if total_assets == 0:
            return False
        
        # Allow for small rounding differences (1% tolerance)
        tolerance = max(total_assets * 0.01, 1000)
        
        return abs(total_assets - total_liab_equity) <= tolerance
    
    @staticmethod
    def _calculate_accuracy(data: Dict[str, Any]) -> float:
        """Calculate data accuracy percentage"""
        # Define key fields that should have values
        key_fields = [
            'current_assets', 'total_assets', 'current_liabilities', 
            'total_liabilities', 'total_equity'
        ]
        
        # Check completeness of key fields
        complete_key_fields = sum(1 for field in key_fields if data.get(field, 0) != 0)
        key_completeness = (complete_key_fields / len(key_fields)) * 100
        
        # Check balance equation
        balance_score = 100 if data.get('balance_check', False) else 50
        
        # Check for reasonable values (no negative assets/equity)
        reasonableness_score = 100
        if data.get('total_assets', 0) < 0:
            reasonableness_score -= 30
        if data.get('total_equity', 0) < 0:
            reasonableness_score -= 20
        
        # Final accuracy score
        accuracy = min((key_completeness + balance_score + reasonableness_score) / 3, 100)
        
        return max(0, accuracy)

class DataImputation:
    """Handle missing data imputation using various methods"""
    
    @staticmethod
    def impute_missing_data(data: Dict[str, Any], method: str = 'knn', 
                           industry_benchmarks: Dict[str, float] = None) -> Dict[str, Any]:
        """Impute missing data using specified method"""
        
        if method == 'knn':
            return DataImputation._knn_imputation(data)
        elif method == 'random_forest':
            return DataImputation._random_forest_imputation(data)
        elif method == 'industry_benchmark':
            return DataImputation._industry_benchmark_imputation(data, industry_benchmarks)
        elif method == 'financial_ratios':
            return DataImputation._ratio_based_imputation(data)
        else:
            return data
    
    @staticmethod
    def _knn_imputation(data: Dict[str, Any]) -> Dict[str, Any]:
        """Use KNN imputation for missing values"""
        try:
            # Convert to DataFrame for easier processing
            df = pd.DataFrame([data])
            
            # Only process numeric columns
            numeric_cols = df.select_dtypes(include=[np.number]).columns
            
            if len(numeric_cols) > 1:
                # Apply KNN imputation
                imputer = KNNImputer(n_neighbors=3)
                df[numeric_cols] = imputer.fit_transform(df[numeric_cols])
                
                # Convert back to dictionary
                return df.iloc[0].to_dict()
            
            return data
        except Exception as e:
            logger.error(f"Error in KNN imputation: {e}")
            return data
    
    @staticmethod
    def _random_forest_imputation(data: Dict[str, Any]) -> Dict[str, Any]:
        """Use Random Forest for imputation"""
        try:
            # This would require historical data for training
            # For now, return original data
            return data
        except Exception as e:
            logger.error(f"Error in Random Forest imputation: {e}")
            return data
    
    @staticmethod
    def _industry_benchmark_imputation(data: Dict[str, Any], 
                                     benchmarks: Dict[str, float]) -> Dict[str, Any]:
        """Use industry benchmarks for imputation"""
        if not benchmarks:
            return data
        
        try:
            for key, value in data.items():
                if (value == 0 or pd.isna(value)) and key in benchmarks:
                    # Use industry benchmark as percentage of total assets
                    if 'total_assets' in data and data['total_assets'] > 0:
                        data[key] = data['total_assets'] * benchmarks[key]
            
            return data
        except Exception as e:
            logger.error(f"Error in industry benchmark imputation: {e}")
            return data
    
    @staticmethod
    def _ratio_based_imputation(data: Dict[str, Any]) -> Dict[str, Any]:
        """Use financial ratios for imputation"""
        try:
            # Example: Estimate accounts receivable based on revenue (if available)
            # This would need to be expanded with more sophisticated ratio relationships
            
            # Common ratios for estimation
            if data.get('accounts_receivable', 0) == 0 and data.get('revenue', 0) > 0:
                # Assume 45-day collection period (industry average)
                data['accounts_receivable'] = data['revenue'] * (45 / 365)
            
            if data.get('inventory', 0) == 0 and data.get('revenue', 0) > 0:
                # Assume inventory turnover of 6 times per year
                data['inventory'] = data['revenue'] / 6
            
            # Estimate cash based on current ratio
            if data.get('cash_and_equivalents', 0) == 0:
                current_liabilities = data.get('current_liabilities', 0)
                if current_liabilities > 0:
                    # Assume 10% of current liabilities as minimum cash
                    data['cash_and_equivalents'] = current_liabilities * 0.1
            
            # Estimate accounts payable based on cost structure
            if data.get('accounts_payable', 0) == 0 and data.get('total_assets', 0) > 0:
                # Assume 8% of total assets as accounts payable
                data['accounts_payable'] = data['total_assets'] * 0.08
            
            return data
        except Exception as e:
            logger.error(f"Error in ratio-based imputation: {e}")
            return data

class ValidationHelper:
    """Validate financial data and calculations"""
    
    @staticmethod
    def validate_financial_data(data: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate financial data for consistency and reasonableness"""
        errors = []
        
        # Check for negative values where they shouldn't exist
        positive_fields = [
            'total_assets', 'current_assets', 'cash_and_equivalents',
            'accounts_receivable', 'inventory'
        ]
        
        for field in positive_fields:
            if data.get(field, 0) < 0:
                errors.append(f"{field} should not be negative")
        
        # Check balance sheet equation
        total_assets = data.get('total_assets', 0)
        total_liabilities = data.get('total_liabilities', 0)
        total_equity = data.get('total_equity', 0)
        
        if total_assets > 0:
            balance_diff = abs(total_assets - (total_liabilities + total_equity))
            if balance_diff > total_assets * 0.05:
                errors.append("Balance sheet equation does not balance (Assets ≠ Liabilities + Equity)")
        
        # Check for unreasonable ratios
        if data.get('current_liabilities', 1) > 0:
            current_ratio = data.get('current_assets', 0) / data.get('current_liabilities', 1)
            if current_ratio > 10:
                errors.append("Current ratio seems unusually high (>10)")
            elif current_ratio < 0.1:
                errors.append("Current ratio seems unusually low (<0.1)")
        
        # Check debt-to-equity ratio
        if data.get('total_equity', 1) > 0:
            debt_to_equity = data.get('total_liabilities', 0) / data.get('total_equity', 1)
            if debt_to_equity > 5:
                errors.append("Debt-to-equity ratio seems unusually high (>5)")
        
        # Check for missing critical values
        critical_fields = ['total_assets', 'total_liabilities', 'total_equity']
        for field in critical_fields:
            if data.get(field, 0) == 0:
                errors.append(f"Critical field {field} is missing or zero")
        
        return len(errors) == 0, errors
    
    @staticmethod
    def calculate_data_quality_score(data: Dict[str, Any], 
                                   validation_errors: List[str]) -> float:
        """Calculate overall data quality score"""
        
        # Base score on completeness
        total_fields = len([k for k in data.keys() if isinstance(data[k], (int, float))])
        complete_fields = len([k for k in data.keys() 
                             if isinstance(data[k], (int, float)) and data[k] != 0])
        
        completeness_score = (complete_fields / total_fields) * 100 if total_fields > 0 else 0
        
        # Deduct points for validation errors
        error_penalty = len(validation_errors) * 10
        
        # Balance sheet equation bonus
        balance_bonus = 10 if data.get('balance_check', False) else 0
        
        # Final score
        quality_score = max(0, min(100, completeness_score - error_penalty + balance_bonus))
        
        return quality_score

class CashFlowProcessor:
    """Process cash flow statement data"""
    
    @staticmethod
    def process_cash_flow_data(data: Dict[str, Any]) -> Dict[str, Any]:
        """Process and validate cash flow data"""
        
        processed_data = {
            'net_income': data.get('net_income', 0),
            'depreciation_and_amortization': data.get('depreciation_and_amortization', 0),
            'stock_based_compensation': data.get('stock_based_compensation', 0),
            'changes_in_working_capital': data.get('changes_in_working_capital', 0),
            'accounts_receivable': data.get('accounts_receivable', 0),
            'inventory': data.get('inventory', 0),
            'accounts_payable': data.get('accounts_payable', 0),
            'net_cash_from_operating_activities': data.get('net_cash_from_operating_activities', 0),
            'capital_expenditures': data.get('capital_expenditures', 0),
            'acquisitions': data.get('acquisitions', 0),
            'net_cash_from_investing_activities': data.get('net_cash_from_investing_activities', 0),
            'dividends_paid': data.get('dividends_paid', 0),
            'share_repurchases': data.get('share_repurchases', 0),
            'net_cash_from_financing_activities': data.get('net_cash_from_financing_activities', 0),
        }
        
        # Calculate derived metrics
        processed_data = CashFlowProcessor._calculate_derived_metrics(processed_data)
        
        # Validate cash flow relationships
        processed_data['validation_errors'] = CashFlowProcessor._validate_cash_flow_data(processed_data)
        
        return processed_data
    
    @staticmethod
    def _calculate_derived_metrics(data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate derived cash flow metrics"""
        
        # Free Cash Flow
        operating_cf = data.get('net_cash_from_operating_activities', 0)
        capex = data.get('capital_expenditures', 0)
        data['free_cash_flow'] = operating_cf - abs(capex)
        
        # OCF to Net Income Ratio
        net_income = data.get('net_income', 1)
        data['ocf_to_net_income_ratio'] = operating_cf / net_income if net_income != 0 else 0
        
        # Net Change in Cash
        investing_cf = data.get('net_cash_from_investing_activities', 0)
        financing_cf = data.get('net_cash_from_financing_activities', 0)
        data['net_change_in_cash'] = operating_cf + investing_cf + financing_cf
        
        return data
    
    @staticmethod
    def _validate_cash_flow_data(data: Dict[str, Any]) -> List[str]:
        """Validate cash flow data for consistency"""
        errors = []
        
        # Check free cash flow calculation
        operating_cf = data.get('net_cash_from_operating_activities', 0)
        capex = data.get('capital_expenditures', 0)
        calculated_fcf = operating_cf - abs(capex)
        reported_fcf = data.get('free_cash_flow', 0)
        
        if abs(calculated_fcf - reported_fcf) > 1000:
            errors.append("Free cash flow calculation inconsistency")
        
        # Check for unreasonable values
        if abs(operating_cf) > 1e12:  # $1 trillion
            errors.append("Operating cash flow seems unreasonably large")
        
        # Check working capital changes consistency
        ar_change = data.get('accounts_receivable', 0)
        inv_change = data.get('inventory', 0)
        ap_change = data.get('accounts_payable', 0)
        total_wc_change = data.get('changes_in_working_capital', 0)
        
        calculated_wc_change = -(ar_change + inv_change) + ap_change
        if abs(calculated_wc_change - total_wc_change) > max(abs(total_wc_change) * 0.2, 10000):
            errors.append("Working capital components don't match total change")
        
        return errors

class FileManager:
    """Manage file operations and storage"""
    
    @staticmethod
    def save_uploaded_file(file, upload_folder: str) -> str:
        """Save uploaded file and return file path"""
        try:
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)
            
            filename = DocumentProcessor.secure_filename_custom(file.filename)
            
            # Add timestamp to avoid conflicts
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_")
            filename = timestamp + filename
            
            file_path = os.path.join(upload_folder, filename)
            file.save(file_path)
            
            return file_path
        except Exception as e:
            logger.error(f"Error saving file: {e}")
            raise
    
    @staticmethod
    def cleanup_temp_files(file_paths: List[str]) -> None:
        """Clean up temporary files"""
        for file_path in file_paths:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    logger.info(f"Cleaned up temp file: {file_path}")
            except Exception as e:
                logger.warning(f"Could not remove temp file {file_path}: {e}")
    
    @staticmethod
    def get_file_info(file_path: str) -> Dict[str, Any]:
        """Get file information"""
        try:
            if os.path.exists(file_path):
                stat = os.stat(file_path)
                return {
                    'size': stat.st_size,
                    'created': datetime.fromtimestamp(stat.st_ctime),
                    'modified': datetime.fromtimestamp(stat.st_mtime),
                    'extension': Path(file_path).suffix.lower()
                }
            return {}
        except Exception as e:
            logger.error(f"Error getting file info: {e}")
            return {}

class IndustryBenchmarks:
    """Handle industry-specific benchmarks and ratios"""
    
    @staticmethod
    def get_industry_benchmarks(industry: str) -> Dict[str, float]:
        """Get industry benchmark ratios"""
        
        # Default benchmarks (as percentage of total assets)
        default_benchmarks = {
            'cash_and_equivalents': 0.05,
            'accounts_receivable': 0.15,
            'inventory': 0.20,
            'current_assets': 0.40,
            'accounts_payable': 0.08,
            'current_liabilities': 0.15,
            'long_term_debt': 0.25,
            'total_equity': 0.50,
            'current_ratio': 2.0,
            'debt_to_equity': 0.5,
            'roa': 0.05,
            'roe': 0.12
        }
        
        # Industry-specific adjustments
        industry_adjustments = {
            'Technology': {
                'cash_and_equivalents': 0.15,
                'inventory': 0.05,
                'accounts_receivable': 0.10,
                'current_ratio': 2.5,
                'debt_to_equity': 0.3,
                'roa': 0.08,
                'roe': 0.15
            },
            'Manufacturing': {
                'inventory': 0.30,
                'property_plant_equipment': 0.40,
                'current_ratio': 1.8,
                'debt_to_equity': 0.7,
                'roa': 0.04,
                'roe': 0.10
            },
            'Retail': {
                'inventory': 0.35,
                'cash_and_equivalents': 0.03,
                'current_ratio': 1.5,
                'debt_to_equity': 0.8,
                'roa': 0.06,
                'roe': 0.13
            },
            'Healthcare': {
                'accounts_receivable': 0.25,
                'inventory': 0.10,
                'current_ratio': 2.2,
                'debt_to_equity': 0.4,
                'roa': 0.07,
                'roe': 0.14
            },
            'Financial Services': {
                'cash_and_equivalents': 0.25,
                'inventory': 0.01,
                'current_ratio': 1.2,
                'debt_to_equity': 3.0,
                'roa': 0.01,
                'roe': 0.12
            },
            'Energy': {
                'property_plant_equipment': 0.60,
                'inventory': 0.15,
                'current_ratio': 1.6,
                'debt_to_equity': 0.9,
                'roa': 0.03,
                'roe': 0.08
            }
        }
        
        benchmarks = default_benchmarks.copy()
        if industry in industry_adjustments:
            benchmarks.update(industry_adjustments[industry])
        
        return benchmarks
    
    @staticmethod
    def get_industry_list() -> List[str]:
        """Get list of supported industries"""
        return [
            'Technology',
            'Manufacturing',
            'Retail',
            'Healthcare',
            'Financial Services',
            'Energy',
            'Real Estate',
            'Transportation',
            'Telecommunications',
            'Consumer Goods',
            'Materials',
            'Utilities'
        ]

def format_currency(amount: float, currency: str = 'USD') -> str:
    """Format currency for display"""
    if pd.isna(amount) or amount == 0:
        return "$0"
    
    if currency == 'USD':
        if abs(amount) >= 1e9:
            return f"${amount/1e9:.1f}B"
        elif abs(amount) >= 1e6:
            return f"${amount/1e6:.1f}M"
        elif abs(amount) >= 1e3:
            return f"${amount/1e3:.1f}K"
        else:
            return f"${amount:,.0f}"
    else:
        return f"{amount:,.2f}"

def calculate_financial_ratios(balance_sheet: Dict[str, Any], 
                             cash_flow: Dict[str, Any] = None) -> Dict[str, float]:
    """Calculate comprehensive financial ratios"""
    
    ratios = {}
    
    # Liquidity Ratios
    current_assets = balance_sheet.get('current_assets', 0)
    current_liabilities = max(balance_sheet.get('current_liabilities', 1), 1)
    inventory = balance_sheet.get('inventory', 0)
    cash = balance_sheet.get('cash_and_equivalents', 0)
    
    ratios['current_ratio'] = current_assets / current_liabilities
    ratios['quick_ratio'] = (current_assets - inventory) / current_liabilities
    ratios['cash_ratio'] = cash / current_liabilities
    
    # Leverage Ratios
    total_debt = (balance_sheet.get('short_term_debt', 0) + 
                 balance_sheet.get('long_term_debt', 0))
    total_equity = max(balance_sheet.get('total_equity', 1), 1)
    total_assets = max(balance_sheet.get('total_assets', 1), 1)
    
    ratios['debt_to_equity'] = total_debt / total_equity
    ratios['debt_to_assets'] = total_debt / total_assets
    ratios['equity_ratio'] = total_equity / total_assets
    
    # Asset Efficiency Ratios
    ratios['asset_utilization'] = current_assets / total_assets
    
    # Cash Flow Ratios (if available)
    if cash_flow:
        operating_cf = cash_flow.get('net_cash_from_operating_activities', 0)
        net_income = max(cash_flow.get('net_income', 1), 1)
        free_cf = cash_flow.get('free_cash_flow', 0)
        
        ratios['ocf_to_net_income'] = operating_cf / net_income
        ratios['operating_cf_ratio'] = operating_cf / current_liabilities
        ratios['free_cf_yield'] = free_cf / total_assets
    
    return ratios

def detect_financial_anomalies(data: Dict[str, Any]) -> List[Dict[str, str]]:
    """Detect anomalies in financial data"""
    
    anomalies = []
    
    # Check for negative equity
    if data.get('total_equity', 0) < 0:
        anomalies.append({
            'type': 'critical',
            'description': 'Negative shareholders equity detected',
            'field': 'total_equity',
            'value': data.get('total_equity', 0)
        })
    
    # Check for excessive debt
    debt_to_equity = calculate_financial_ratios(data).get('debt_to_equity', 0)
    if debt_to_equity > 3:
        anomalies.append({
            'type': 'warning',
            'description': f'Very high debt-to-equity ratio: {debt_to_equity:.2f}',
            'field': 'debt_to_equity',
            'value': debt_to_equity
        })
    
    # Check for low liquidity
    current_ratio = calculate_financial_ratios(data).get('current_ratio', 0)
    if current_ratio < 0.5:
        anomalies.append({
            'type': 'critical',
            'description': f'Very low current ratio: {current_ratio:.2f}',
            'field': 'current_ratio',
            'value': current_ratio
        })
    
    # Check for unrealistic cash position
    cash_ratio = calculate_financial_ratios(data).get('cash_ratio', 0)
    if cash_ratio > 1:
        anomalies.append({
            'type': 'info',
            'description': f'Unusually high cash position: {cash_ratio:.2f}',
            'field': 'cash_ratio',
            'value': cash_ratio
        })
    
    return anomalies

def get_industry_benchmarks(industry: str) -> Dict[str, float]:
    """Get industry benchmark ratios - Wrapper function"""
    return IndustryBenchmarks.get_industry_benchmarks(industry)

def log_processing_step(step: str, details: str = "") -> None:
    """Log processing steps for debugging"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logger.info(f"[{timestamp}] {step}: {details}")

def generate_processing_report(data: Dict[str, Any], 
                             validation_errors: List[str],
                             anomalies: List[Dict[str, str]]) -> Dict[str, Any]:
    """Generate comprehensive processing report"""
    
    report = {
        'timestamp': datetime.now().isoformat(),
        'data_quality_score': ValidationHelper.calculate_data_quality_score(data, validation_errors),
        'validation_errors': validation_errors,
        'anomalies': anomalies,
        'financial_ratios': calculate_financial_ratios(data),
        'completeness': {
            'total_fields': len(data),
            'complete_fields': len([k for k, v in data.items() if v != 0]),
            'missing_fields': [k for k, v in data.items() if v == 0]
        },
        'balance_check': data.get('balance_check', False),
        'accuracy_percentage': data.get('accuracy_percentage', 0)
    }
    
    return report

def process_uploaded_document(file_path: str, document_type: str = None) -> Dict[str, Any]:
    """Main function to process uploaded financial documents"""
    
    try:
        log_processing_step("Starting document processing", f"File: {file_path}")
        
        # Determine file type
        file_extension = Path(file_path).suffix.lower()
        
        # Extract text and tables based on file type
        if file_extension == '.pdf':
            text = DocumentProcessor.extract_text_from_pdf(file_path)
            tables = DocumentProcessor.extract_tables_from_pdf(file_path)
        elif file_extension in ['.xlsx', '.xls']:
            excel_data = DocumentProcessor.process_excel_file(file_path)
            text = ""
            tables = list(excel_data.values())
        elif file_extension == '.csv':
            csv_data = DocumentProcessor.process_csv_file(file_path)
            text = ""
            tables = [csv_data] if not csv_data.empty else []
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")
        
        # Identify document type if not provided
        if not document_type:
            document_type = FinancialDataExtractor.identify_statement_type(text)
        
        log_processing_step("Document type identified", document_type)
        
        # Extract financial data based on document type
        if document_type == 'balance_sheet':
            extracted_data = FinancialDataExtractor.extract_balance_sheet_data(text, tables)
            processed_data = BalanceSheetProcessor.process_balance_sheet_data(extracted_data)
        elif document_type == 'cash_flow':
            extracted_data = FinancialDataExtractor.extract_cash_flow_data(text, tables)
            processed_data = CashFlowProcessor.process_cash_flow_data(extracted_data)
        else:
            # For income statements or unknown documents, extract basic financial values
            extracted_data = FinancialDataExtractor.extract_financial_values(text)
            processed_data = extracted_data
        
        # Add metadata
        processed_data.update({
            'year': FinancialDataExtractor.extract_year_from_text(text),
            'generated_at': datetime.now(),
            'data_source': 'uploaded_documents',
            'document_type': document_type,
            'file_name': Path(file_path).name
        })
        
        # Validate data
        is_valid, validation_errors = ValidationHelper.validate_financial_data(processed_data)
        processed_data['validation_errors'] = validation_errors
        
        # Detect anomalies
        anomalies = detect_financial_anomalies(processed_data)
        processed_data['anomalies'] = anomalies
        
        # Generate processing report
        report = generate_processing_report(processed_data, validation_errors, anomalies)
        processed_data['processing_report'] = report
        
        log_processing_step("Document processing completed", f"Quality Score: {report['data_quality_score']:.1f}%")
        
        return {
            'success': True,
            'data': processed_data,
            'document_type': document_type,
            'report': report
        }
        
    except Exception as e:
        logger.error(f"Error processing document: {e}")
        return {
            'success': False,
            'error': str(e),
            'data': None
        }

# Global configuration
CONFIG = {
    'max_file_size': 50 * 1024 * 1024,  # 50MB
    'allowed_extensions': DocumentProcessor.ALLOWED_EXTENSIONS,
    'temp_folder': 'temp',
    'upload_folder': 'uploads',
    'default_industry': 'Technology',
    'validation_tolerance': 0.05,  # 5% tolerance for validation
    'min_accuracy_threshold': 70.0  # Minimum acceptable accuracy percentage
}

# Initialize global objects
document_processor = DocumentProcessor()
financial_extractor = FinancialDataExtractor()
balance_sheet_processor = BalanceSheetProcessor()
data_imputation = DataImputation()
validation_helper = ValidationHelper()
cash_flow_processor = CashFlowProcessor()
file_manager = FileManager()
industry_benchmarks = IndustryBenchmarks()